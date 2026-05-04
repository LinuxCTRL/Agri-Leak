import logging
import openpyxl
from pathlib import Path
import sys

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
)
logger = logging.getLogger(__name__)


def cleanup_tonnage_sheets(file_path: str):
    """Remove all sheets except 'SUIVI JOUR' and 'QNZ N°' from an Excel file."""
    logger.info(f"Processing: {file_path}")
    
    wb = openpyxl.load_workbook(file_path)
    original_sheets = wb.sheetnames
    logger.info(f"  Found {len(original_sheets)} sheets: {original_sheets}")
    
    sheets_to_keep = [s for s in wb.sheetnames if 'SUIVI JOUR' in s or 'QNZ N°' in s]
    
    if not sheets_to_keep:
        logger.warning(f"  SKIP: no matching sheets ('SUIVI JOUR' or 'QNZ N°')")
        return False
    
    sheets_removed = [s for s in wb.sheetnames if s not in sheets_to_keep]
    logger.info(f"  Removing {len(sheets_removed)} sheets: {sheets_removed}")
    logger.info(f"  Keeping {len(sheets_to_keep)} sheets: {sheets_to_keep}")
    
    for s in sheets_removed:
        del wb[s]
    
    wb.save(file_path)
    logger.info(f"  Done: saved to {file_path}")
    return True


def main():
    logger.info("Starting Excel cleanup script")
    
    data_dir = Path("data")
    files = sorted([f for f in data_dir.iterdir() if f.name.startswith("TONNAGE_QNZ")])
    
    if not files:
        logger.warning("No TONNAGE_QNZ_* files found in data/")
        return
    
    logger.info(f"Found {len(files)} files to process")
    
    for file in files:
        try:
            cleanup_tonnage_sheets(str(file))
        except Exception as e:
            logger.error(f"ERROR {file.name}: {e}")
    
    logger.info("Excel cleanup complete")


if __name__ == "__main__":
    main()
    sys.exit(0)